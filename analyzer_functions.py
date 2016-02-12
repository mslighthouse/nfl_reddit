import sqlite3
import operator
import datetime, calendar
from collections import defaultdict
from openpyxl import Workbook, load_workbook

# Created by Maxwell Smith - Running Python 2.7.10
# This file holds the function definitions for every function used in the analyzer.py file.
# This allows for more readability in the analyzer.py file as well as clarity as to what functions
# are doing what.

# connect to sqlite server
conn = sqlite3.connect('nfl.db')
c = conn.cursor()

# Sets up the Excel workbook sheet
# Returns nothing
def setup_workbook(database):
    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = str(database)

    # create titles
    ws['A1'] = "Username"
    ws['B1'] = "Post Count"
    ws['C1'] = "Flair"
    ws['E1'] = "Word"
    ws['F1'] = "Times counted"
    ws['G1'] = "Time"
    ws['G2'] = "Posts per Minute"
    
    # Save
    wb.save(database + '.xlsx')

# Analyzes the unique users and unique user post count. Places them into spreadsheet
# Returns nothing
def unique_users(database):
    # Reopen Workbook
    wb = wb = load_workbook(filename = database + '.xlsx')
    ws = wb.active
    
    prev_user  = ""
    user_total = 0
    post_count = 0
    delete_num = 0
    execute    = 'SELECT author FROM ' + database + ' ORDER BY author'

    for row in c.execute(execute):
        if row[0] != prev_user and row[0] != "None":
            prev_user = row[0]
            user_total += 1
        if row[0] == "None":
            delete_num += 1
        post_count += 1
    
    print("There were " + str(user_total) + " individual users contributing " + str(post_count) + " total comments.")
    print("Of the " + str(post_count) + " total posts, " + str(delete_num) + " have been deleted.\n")

    rown = 2 # row number iterator
    # Users and their post amount
    for row in c.execute('SELECT author, count(*) FROM ' + database + ' GROUP BY author ORDER BY author'):
        ws.cell(row=rown, column=1).value = row[0]
        ws.cell(row=rown, column=2).value = int(float(row[1]))
        rown += 1

    # Resave workbook
    wb.save(database + '.xlsx')

def flairs(database):
    # Reopen workbook
    wb = wb = load_workbook(filename = database + '.xlsx')
    ws = wb.active
    
    rown = 2
    execute = 'SELECT flair FROM ' + database + ' GROUP BY author ORDER BY author'
    for row in c.execute (execute):
        
        flair = row[0]
        
        ws.cell(row=rown, column=3).value = flair
        rown = rown + 1
    
    # Resave Workbook
    wb.save(database + '.xlsx')

def gilded(database):
    gild = 0
    for row in c.execute('SELECT gilded FROM ' + database):
        if (row[0] == "TRUE"):
            gild += 1

    print(str(gild) + " comments were given gold. That means the superbowl threads generated $" + str(3.99 * gild) + " through gildiing.")

def comments(database):
    # Reopen Workbook
    wb = load_workbook(filename = database + '.xlsx')
    ws = wb.active
    
    # Comment bodies and individual word frequency
    dict = defaultdict(int)

    for row in c.execute ('SELECT comment FROM ' + database):
        comment = row[0].lower()
        # Do stuff here
        
        comment = comment.split()
        for word in comment:
            word = word.lower()
            dict[word] += 1

    # Sort and organize dictionary for placement in Worksheet
    sorted_dict = sorted(dict.items(), key=operator.itemgetter(1))

    rown = 2
    for item in sorted_dict:
        ws.cell(row=rown, column=5).value = item[0]
        ws.cell(row=rown, column=6).value = item[1]
        rown += 1

    # Resave Workbook
    wb.save(database + '.xlsx')

def comments_per_minute(database):
    # Reopen Workbook
    wb = wb = load_workbook(filename = database + '.xlsx')
    ws = wb.active

    comdict = defaultdict(int)
    for row in c.execute('SELECT time_utc FROM ' + database):
        date_time = datetime.datetime.utcfromtimestamp(int(row[0]))
        fmt_time  = date_time.strftime("%D %H:%M")
        fmt_time = str(fmt_time)
        comdict[fmt_time] += 1

    sorted_dict = sorted(comdict.items(), key=operator.itemgetter(1))
    rown=2
    for item in sorted_dict:
        ws.cell(row=rown, column=7).value = item[0]
        ws.cell(row=rown, column=8).value = item[1]
        rown += 1


    # Save Workbook
    wb.save(database + '.xlsx')
